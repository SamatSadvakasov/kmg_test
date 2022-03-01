from openpyxl import load_workbook
from datetime import datetime, timedelta

wb = load_workbook(filename='task/AA_111/01.07.2020.xlsx')
sheets = wb.sheetnames
ws = wb[sheets[0]]
total_amount_sec = 1*24*60*60  # 86400 seconds
# Time in the beginning of cycle
file_datetime = '2020-07-01 00:00:00'
FMT = '%Y-%m-%d %H:%M:%S'
# Converting str to datetime.datetime
file_datetime = datetime.strptime(file_datetime, FMT)
all_day_res = list()
for row in ws.iter_rows(min_row=2):
    # First column of table (datetime.time)
    start_time = row[0].value
    start_time_in_sec = start_time.hour*60*60 + start_time.minute*60 + start_time.second

    # Second column of table (float or x)
    value = row[1].value
    # Third column of table (datetime.time)
    duration = row[2].value
    duration_in_sec = duration.hour*60*60 + duration.minute*60 + duration.second

    if value == 'x':
        value = 0

    for i in range(duration_in_sec):
        all_day_res.append([file_datetime + timedelta(seconds=start_time_in_sec), value, duration])
        start_time_in_sec += 1


# From Start of Day (00:00:00) To First Time in File
time_diff = all_day_res[0][0] - file_datetime
zero_to_first = list()
for i in range(time_diff.seconds):
    zero_to_first.append([file_datetime + timedelta(seconds=i), 0, file_datetime.time()])

# From Last Record To End of Day (23:59:59)
time_diff = (file_datetime + timedelta(seconds=total_amount_sec-1)) - all_day_res[-1][0]
last_to_end = list()
for i in range(time_diff.seconds):
    last_to_end.append([all_day_res[-1][0] + timedelta(seconds=i+1), 0, file_datetime.time()])

final_res = [*zero_to_first, *all_day_res, *last_to_end]
if len(final_res) == total_amount_sec:
    print("OK")
