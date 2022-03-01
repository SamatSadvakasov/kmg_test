from openpyxl import load_workbook
from datetime import timedelta, datetime
from database.db import add_field, save_field_data
from database.models import FieldData
import os


def parse_excel(file_path, file_datetime):
    wb = load_workbook(filename=file_path)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    total_amount_sec = 1 * 24 * 60 * 60  # 86400 seconds

    all_day_res = list()
    for row in ws.iter_rows(min_row=2):
        # First column of table (datetime.time)
        start_time = row[0].value
        start_time_in_sec = start_time.hour * 60 * 60 + start_time.minute * 60 + start_time.second

        # Second column of table (float or x)
        value = row[1].value
        # Third column of table (datetime.time)
        duration = row[2].value
        duration_in_sec = duration.hour * 60 * 60 + duration.minute * 60 + duration.second

        if value == 'x':
            value = 0

        for i in range(duration_in_sec):
            all_day_res.append([file_datetime + timedelta(seconds=start_time_in_sec), value, duration])
            start_time_in_sec += 1

    # From Start of Day (00:00:00) To First Time in File
    time_diff = all_day_res[0][0] - file_datetime
    zero_to_first = list()
    for i in range(time_diff.seconds):
        zero_to_first.append([file_datetime + timedelta(seconds=i), 0, file_datetime.time()])

    # From Last Record To End of Day (23:59:59)
    time_diff = (file_datetime + timedelta(seconds=total_amount_sec - 1)) - all_day_res[-1][0]
    last_to_end = list()
    for i in range(time_diff.seconds):
        last_to_end.append([all_day_res[-1][0] + timedelta(seconds=i + 1), 0, file_datetime.time()])

    final_res = [*zero_to_first, *all_day_res, *last_to_end]
    if len(final_res) == total_amount_sec:
        print("OK")
        return final_res
    else:
        print('Something Wrong with Excel File')
        return []


# Format of date in file name
def parsing_folder(folder_path):
    fmt = '%d.%m.%Y'
    instance = []
    for filial in os.listdir(folder_path):
        field_id = add_field(filial)
        for file in os.listdir(folder_path + filial):
            file_datetime = datetime.strptime(file[:-5], fmt)
            file_path = folder_path + filial + '/' + file
            data = parse_excel(file_path, file_datetime)
            for item in data:
                instance.append(FieldData(start_datetime=item[0], value=item[1], duration=item[2], field_id=field_id))

    return save_field_data(instance)
